from openpyxl import load_workbook
from openpyxl import Workbook
# Library for finding path
import os

# Loading source file
wb = load_workbook('input.xlsx')
#Sep = [0], Oct = [1], Nov = [2], ...
ws = wb.worksheets[0]

# Loading basic format of donation receipt
wb4 = load_workbook('DonationReceiptBasicFormat.xlsx')
ws4 = wb4.worksheets[0]

# Getting rid of previous result if there is one
if os.path.exists("DonationReceiptBasicFormat_Result.xlsx"):
    os.remove("DonationReceiptBasicFormat_Result.xlsx")

# Saving basic format to result
wb4.save('DonationReceiptBasicFormat_Result.xlsx')

# Loading the result
wb1 = load_workbook('DonationReceiptBasicFormat_Result.xlsx')
ws1 = wb1.worksheets[0]


#Helper function to iterate all the information and fill up to result file
def copyValue(cMinRow, cMinCol, cMaxCol, pMinRow, pMinCol, pMaxCol):
    x = []
    for cols in ws.iter_rows(min_row = cMinRow, max_row = ws.max_row, min_col = cMinCol, max_col = cMaxCol):
        for cell in cols:
            x.append(cell.value)
    i=0
    for cols in ws1.iter_rows(min_row = pMinRow, max_row = ws.max_row, min_col = pMinCol, max_col = pMaxCol):
        for cell in cols:
            cell.value = x[i]
            i += 1


#ex) 3, 6, 6, 3, 3, 3?
# first 3 : Starting from the 3rd row basic format file
# second 6, third 6 : Fix column of basic format file to be 6, min_col = 6, max_col = 6
# fourth 3 : Starting from the 3rd row of result file
# fifth 3, sixth 3 : Fix column of result file to be 6, min_col = 6, max_col = 6
copyValue(3, 6, 6, 3, 3, 3)

copyValue(3, 7, 7, 3, 4, 4)

copyValue(3, 5, 5, 3, 5, 5)


date = ''
if ws.title[5:] == '9월' or ws.title[5:] == '4월' or ws.title[5:] == '6월' or ws.title[5:] == '11월':
        date = '30'
else:
    date = '31'
    if ws.title[5:] == '2월':
        date = '28'

month = ''
if len(ws.title) == 7:
    month = ws.title[5:6].zfill(2)
else:
    month = ws.title[5:7]

year = ws.title[:4]

# Iterate Donation Date and Receipt Date
# For Sep 2018, '2018.09.30', Oct 2018, '2018.10.31', Nov 2018 '2018.11.30.', ...
for cols in ws1.iter_rows(min_row = 3, min_col = 6, max_col = 7):
    for cell in cols:
        cell.value = ws.title[:4] + '.' + month + '.' + date

# Iterate years
i = 0
for cols in ws1.iter_rows(min_row = 3, min_col = 1, max_col = 1):
    for cell in cols:
        #cell.value = '201809-bbq-' + str(i).zfill(4)
        cell.value = year + month + '-bbq-' + str(i).zfill(4)
        i += 1

# Iternate name of the person and the business
x0 = []
x1 = []
x2 = []
for cols in ws.iter_rows(min_row = 3, min_col = 2, max_col = 2):
    for cell in cols:
        x1.append(cell.value)
for cols in ws.iter_rows(min_row = 3, min_col = 3, max_col = 3):
    for cell in cols:
        x2.append(cell.value)


# Concat x1 and x2
# Make tuple with busienss and its owner
x3 = zip(x1, x2)

for x in x3:
    x0.append(x)

x4 = []
for i in range(len(x0)):
    x4.append(' '.join(x0[i]))
    i += 1

# Paste into Result file
i=0
for cols in ws1.iter_rows(min_row = 3, min_col = 2, max_col = 2):
    for cell in cols:
        cell.value = x4[i]
        i += 1

# Make error.txt file to collect all 'N/A'
if os.path.exists("error.txt"):
    os.remove("error.txt")

# If there is '#N/A' add it on error.txt and erase that line from the input
datafile = open('error.txt', 'a')

for cols in ws1.iter_cols(min_row = 3, min_col = 2, max_col = 2):
    for cell in cols:
        if '#N/A' in cell.value:
            datafile.write(str(ws1['A' + str(cell.row)].value[11:])
            + ' ' + cell.value.split(' ')[0]+ '\n')
            ws1.delete_rows(cell.row,1)
            i += 1

datafile.write('\n')

x=[]
for cols in ws1.iter_rows(min_row = 3, min_col = 5, max_col = 5):
    for cell in cols:
        if cell.value <= 0:
                datafile.write(str(ws1['A' + str(cell.row)].value[11:])
                + ' ' + ws1['B' + str(cell.row)].value.split(' ')[0]+ '\n')
                x.append(str(ws1['A' + str(cell.row)].value[11:]))

for cols in ws1.iter_cols(min_row = 3, min_col = 1, max_col = 1):
    for cell in cols:
        if any(cell.value[11:] == num for num in x):
            ws1.delete_rows(cell.row, 1)

datafile.close()

# Save it to result file
wb1.save('DonationReceiptBasicFormat_Result.xlsx')