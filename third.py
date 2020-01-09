from openpyxl import load_workbook
# Library for finding path
import os
# Library for erasing non-empty folders
import shutil

# Loading source file
wb = load_workbook('input.xlsx')
#Sep = [0], Oct = [1], Nov = [2], ...
ws = wb.worksheets[1]

# Loading basic format of donation receipt
wb6 = load_workbook('IndividualReceipt.xlsx')
ws6 = wb6.worksheets[0]

dir_name = 'Individual(' + ws.title[5:] + ')'

# Getting rid of previous result if there is one
if os.path.isdir(dir_name):
    shutil.rmtree(dir_name)

# Remake directory
os.makedirs(dir_name)

# Iterall the whole file
for i in range(ws.max_row-2):
    if ws['C' + str(i+3)].value == '#N/A':
        continue

    if ws['E' + str(i+3)].value <= 0:
        continue

    if ws.title[5:7] == str(10) or ws.title[5:7] == str(11) or ws.title[5:7] == str(12):
        file_name = ws.title[:4] + ws.title[5:7] + '-bbq-' + str(i).zfill(4) + '(' + str(ws['F' + str(i+3)].value) + ').xlsx'
    else:
        file_name = ws.title[:4] + ws.title[5:6].zfill(2) + '-bbq-' + str(i).zfill(4) + '(' + str(ws['F' + str(i+3)].value) + ').xlsx'

    wb6.save(dir_name + '/' + file_name)
    wb3 = load_workbook(dir_name + '/' + file_name)
    ws3 = wb3.worksheets[0]

    ws3.title = ws.title
    ws3['B7'] = ws['C' + str(i+3)].value
    ws3['E10'] = ws['F' + str(i+3)].value
    ws3['B10'] = ws['B' + str(i+3)].value
    ws3['B11'] = ws['G' + str(i+3)].value
    
    month = 0
    if len(ws.title) == 7:
        month = ws.title[5:6].zfill(2)
    else:
        month = ws.title[5:7]

    ws3['D17'] = ws.title[:4] + '.' + str(month)
    ws3['F17'] = ws['E' + str(i+3)].value
    ws3['D19'] = ws.title[:5]
    ws3['E19'] = ws.title[5:]

    if ws.title[5:] == '9월' or ws.title[5:] == '4월' or ws.title[5:] == '6월' or ws.title[5:] == '11월':
        ws3['F19'] = '30일'
    else:
        ws3['F19'] = '31일'
        if ws.title[5:] == '2월':
            ws3['F19'] = '28일'


    ws3['D20'] = ws['C' + str(i+3)].value
    wb3.create_sheet('총액')
    
    sum_sheet = wb3.worksheets[1]
    sum_value = 0
    number = 1
    for j in range(12):
        sum_sheet['A' + str(j+1)] = str(j+1) + '월'
        sum_sheet['B' + str(j+1)] = 0
        if ws3['E19'].value == sum_sheet['A' + str(j+1)].value:
            sum_sheet['B' + str(j+1)] = ws3['F17'].value

        for root, dirs, files in os.walk('개인별 영수증(' + str(j+1) + '월)'):
            for name in files:
                if str(ws['F' + str(i+3)].value) in name and ws.title[5:] != (str(j+1) + '월'):
                    #print('있음' + str(j+1) + '월')
                    number += 1
                    #print(name)
                    wb7 = load_workbook('개인별 영수증(' + str(j+1) + '월)/' + name)
                    ws7 = wb7.worksheets[0]
                    #print(ws7['F17'].value)
                    sum_sheet['B' + str(j+1)] = ws7['F17'].value
  
        sum_value = sum_value + sum_sheet['B' + str(j+1)].value
        sum_sheet['B13'] = sum_value

    sum_sheet['A13'] = '총액'
    
# Save it to result file
    wb3.save(dir_name + '/' + file_name)