from openpyxl import load_workbook
from openpyxl import Workbook
# Library for finding path
import os

# Loading source file
wb = load_workbook('input.xlsx')
#Sep = [0], Oct = [1], Nov = [2], ...
ws = wb.worksheets[1]

# Loading basic format of donation receipt
wb5 = load_workbook('DonationReceipt_Basic.xlsx')
ws5 = wb5.worksheets[0]

# Getting rid of previous result if there is one
if os.path.exists("DonationReceipt_Result.xlsx"):
    os.remove("DonationReceipt_Result.xlsx")

# Saving basic format to result
wb5.save('DonationReceipt_Result.xlsx')

# Loading the result
wb2 = load_workbook('DonationReceipt_Result.xlsx')
ws2 = wb2.worksheets[0]

#Helper function to iterate all the information and fill up to result file
def copyValue1(cMinRow, cMinCol, cMaxCol, pMinRow, pMinCol, pMaxCol):
    x = []
    for cols in ws.iter_rows(min_row = cMinRow, max_row = ws.max_row * 2 + 3, min_col = cMinCol, max_col = cMaxCol):
        for cell in cols:
            x.append(cell.value)
    i=0
    for cols in ws2.iter_rows(min_row = pMinRow, max_row = ws.max_row * 2 + 3, min_col = pMinCol, max_col = pMaxCol):
        for cell in cols:
            if cell.row % 2 == 0:
                continue
            cell.value = x[i]
            i += 1

# Merge names
for rows in ws2.iter_rows(min_row=9, max_row = ws.max_row * 2 + 3, min_col = 3, max_col = 3):
    for cell in rows:
        if cell.row % 2 == 1:
            continue
        ws2.merge_cells(start_row=cell.row, end_row=cell.row, start_column=3, end_column=4)

# Iterate business number
i = 0
for rows in ws2.iter_rows(min_row = 9, max_row = ws.max_row * 2 + 3, min_col = 1, max_col = 1):
    for cell in rows:
        if cell.row % 2 == 0:
            continue
        cell.value = i
        i += 1

date = ''
if ws.title[5:] == '9월' or ws.title[5:] == '4월' or ws.title[5:] == '6월' or ws.title[5:] == '11월':
        date = '30'
else:
    date = '31'
    if ws.title[5:] == '2월':
        date = '28'

month = ''
if len(ws.title) == 7:
    month = ws.title[5:6].zfill(2)
else:
    month = ws.title[5:7]

year = ws.title[:4]

# Iterate Date of Donation
# For Sep 2018, '2018.09.30', Oct 2018, '2018.10.31', Nov 2018 '2018.11.30.', ...
for rows in ws2.iter_rows(min_row = 9, max_row = ws.max_row * 2 + 3, min_col = 2, max_col = 2):
    for cell in rows:
        if cell.row % 2 == 0:
            continue
        #cell.value = '2018.10.31'
        cell.value = year + '.' + month + '.' + date


# Iterate Date of Receipt(1)
# For Sep 2018, '2018.09.30', Oct 2018, '2018.10.31', Nov 2018 '2018.11.30.', ...
for rows in ws2.iter_rows(min_row = 9, max_row = ws.max_row * 2 + 3, min_col = 9, max_col = 9):
    for cell in rows:
        if cell.row % 2 == 0:
            continue
        #cell.value = '2018.10.31'
        cell.value = year + '.' + month + '.' + date

# Iterate Date of Receipt(2)
i = 0
for cols in ws2.iter_rows(min_row = 9, max_row = ws.max_row * 2 + 3, min_col = 8, max_col = 8):
    for cell in cols:
        if cell.row % 2 == 0:
            continue
        #cell.value = '201810-bbq-' + str(i).zfill(4)
        cell.value = year + month + '-bbq-' + str(i).zfill(4)
        i += 1

# Iterate Address
x = []
for cols in ws.iter_rows(min_row = 3, max_row = ws.max_row * 2 + 3, min_col = 7, max_col = 7):
    for cell in cols:
        x.append(cell.value)
i=0
for cols in ws2.iter_rows(min_row = 10, max_row = ws.max_row * 2 + 3, min_col = 3, max_col = 3):
    for cell in cols:
        if cell.row % 2 == 1:
            continue
        cell.value = x[i]
        i += 1


# Iterate Donator
copyValue1(3, 3, 3, 9, 3, 3)

# Iterate Business number
copyValue1(3, 6, 6, 9, 4, 4)

# Iterate Amount
copyValue1(3, 5, 5, 9, 7, 7)

# Iterate Business name
copyValue1(3, 6, 6, 9, 4, 4)

# If there is '#N/A' add it on error.txt and erase that line from the input
for cols in ws2.iter_cols(min_row = 9, min_col = 3, max_col = 3):
    for cell in cols:
        if cell.value == '#N/A':
            ws2.delete_rows(cell.row,1)

# Save it to result file
wb2.save('DonationReceipt_Result.xlsx')